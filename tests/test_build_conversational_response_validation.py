from __future__ import annotations

import csv
import gzip
import tempfile
import unittest
from pathlib import Path

from src.build_conversational_response_validation import (
    LABEL_COLUMNS,
    audit_context_mismatches,
    manual_label_audit,
    prepare_manual_review,
    validate_human_review,
)


def write_csv(path: Path, rows: list[dict[str, str]], compressed: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle = (
        gzip.open(path, "wt", encoding="utf-8", newline="")
        if compressed
        else path.open("w", encoding="utf-8", newline="")
    )
    with handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


class ConversationalResponseValidationTests(unittest.TestCase):
    def test_context_mismatch_is_traced_to_earlier_caregiver(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            raw = root / "raw/Brown/Adam/a.cha"
            raw.parent.mkdir(parents=True)
            raw.write_text(
                "@UTF8\n*MOT:\told context .\n*CHI:\tintervening .\n*MOT:\tnew context .\n*CHI:\ttarget .\n",
                encoding="utf-8",
            )
            flags = root / "flags.csv.gz"
            write_csv(
                flags,
                [
                    {
                        "dataset": "Brown",
                        "child_id": "Adam",
                        "file": "Adam/a.cha",
                        "line_no": "5",
                        "utt_id": "1",
                        "context_k1": "old context.",
                        "previous_main_utterance_clean": "new context.",
                        "primary_responsive_turn_eligible": "1",
                        "context_k1_matches_nearest_caretaker": "0",
                    }
                ],
                compressed=True,
            )
            output = root / "out"
            audit = audit_context_mismatches(flags, root / "raw", output)
            self.assertEqual(audit["status"], "REVIEW_UNEXPECTED_MISMATCHES")
            self.assertEqual(audit["counts"]["context_points_to_earlier_allowed_caregiver"], 1)
            self.assertEqual(audit["counts"]["unexpected_mismatch_rows"], 1)
            with gzip.open(output / "context_k1_mismatch_rows.csv.gz", "rt", encoding="utf-8") as handle:
                row = next(csv.DictReader(handle))
            self.assertEqual(row["matched_prior_line_no"], "2")
            self.assertEqual(row["matched_prior_caregiver_rank"], "2")

    def test_empty_immediate_caregiver_tier_is_an_expected_scorer_skip(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            raw = root / "raw/Brown/Adam/a.cha"
            raw.parent.mkdir(parents=True)
            raw.write_text(
                "@UTF8\n*MOT:\told context .\n*MOT:\t&=laughs .\n*CHI:\ttarget .\n",
                encoding="utf-8",
            )
            flags = root / "flags.csv.gz"
            write_csv(
                flags,
                [
                    {
                        "dataset": "Brown",
                        "child_id": "Adam",
                        "file": "Adam/a.cha",
                        "line_no": "4",
                        "utt_id": "1",
                        "context_k1": "old context.",
                        "previous_main_utterance_clean": "",
                        "primary_responsive_turn_eligible": "1",
                        "context_k1_matches_nearest_caretaker": "0",
                    }
                ],
                compressed=True,
            )
            output = root / "out"
            audit = audit_context_mismatches(flags, root / "raw", output)
            self.assertEqual(audit["status"], "PASS_ADJUDICATED_EXPECTED_EMPTY_TIERS")
            self.assertEqual(audit["counts"]["adjudicated_expected_rows"], 1)
            self.assertEqual(audit["counts"]["unexpected_mismatch_rows"], 0)
            with gzip.open(
                output / "context_k1_mismatch_rows.csv.gz", "rt", encoding="utf-8"
            ) as handle:
                row = next(csv.DictReader(handle))
            self.assertEqual(row["classification"], "expected_empty_immediate_caregiver_skipped")
            self.assertEqual(row["immediate_previous_raw"], "&=laughs .")
            self.assertEqual(row["context_k1_matches_nearest_nonempty_caregiver"], "1")

    def test_manual_review_is_hash_bound_and_stays_unlabeled(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "manual.csv"
            row = {
                "dataset": "Brown",
                "child_id": "Adam",
                "file": "Adam/a.cha",
                "line_no": "5",
                "utt_id": "1",
                "child_utterance_raw": "target .\u0015",
                "previous_main_utterance_clean": "prompt.",
                "next_main_utterance_clean": "reply.",
                **{column: "" for column in LABEL_COLUMNS},
                "manual_notes": "",
            }
            write_csv(source, [row])
            audit = prepare_manual_review(source, root / "raw", root / "out", expected_rows=1)
            self.assertEqual(audit["status"], "READY_FOR_HUMAN_REVIEW")
            self.assertEqual(audit["rows_with_all_required_labels"], 0)
            self.assertTrue((root / "out/response_function_manual_review.xlsx").is_file())
            with (root / "out/response_function_manual_review.csv").open(encoding="utf-8") as handle:
                reviewed = next(csv.DictReader(handle))
            self.assertTrue(reviewed["review_id"].startswith("RV-"))
            self.assertEqual(len(reviewed["source_row_sha256"]), 64)

    def test_label_gate_requires_every_valid_label(self) -> None:
        base = {
            "dataset": "Brown",
            "child_id": "Adam",
            "file": "a.cha",
            "line_no": "1",
            "utt_id": "1",
            **{column: "1" for column in LABEL_COLUMNS},
        }
        self.assertEqual(manual_label_audit([base], expected_rows=1)["status"], "PASS_HUMAN_LABELS_COMPLETE")
        incomplete = {**base, LABEL_COLUMNS[0]: ""}
        self.assertEqual(manual_label_audit([incomplete], expected_rows=1)["status"], "IN_PROGRESS")
        invalid = {**base, LABEL_COLUMNS[0]: "maybe"}
        self.assertEqual(manual_label_audit([invalid], expected_rows=1)["status"], "IN_PROGRESS")
        self.assertEqual(len(manual_label_audit([invalid], expected_rows=1)["invalid_labels"]), 1)

    def test_human_workbook_can_be_validated_without_overwriting_it(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "manual.csv"
            row = {
                "dataset": "Brown",
                "child_id": "Adam",
                "file": "Adam/a.cha",
                "line_no": "5",
                "utt_id": "1",
                "child_utterance_raw": "target .",
                "previous_main_utterance_clean": "prompt.",
                "next_main_utterance_clean": "reply.",
                **{column: "" for column in LABEL_COLUMNS},
                "manual_notes": "",
            }
            write_csv(source, [row])
            output = root / "out"
            prepare_manual_review(source, root / "raw", output, expected_rows=1)
            workbook_path = output / "response_function_manual_review.xlsx"
            workbook = load_workbook(workbook_path)
            worksheet = workbook["Review"]
            header = [cell.value for cell in worksheet[1]]
            for column in LABEL_COLUMNS:
                worksheet.cell(row=2, column=header.index(column) + 1, value="1")
            workbook.save(workbook_path)
            audit = validate_human_review(workbook_path, source, output, expected_rows=1)
            self.assertEqual(audit["status"], "PASS_HUMAN_LABELS_COMPLETE")
            self.assertTrue(audit["source_binding_passed"])

    def test_human_review_rejects_a_changed_source_hash(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "manual.csv"
            row = {
                "dataset": "Brown",
                "child_id": "Adam",
                "file": "Adam/a.cha",
                "line_no": "5",
                "utt_id": "1",
                **{column: "1" for column in LABEL_COLUMNS},
            }
            write_csv(source, [row])
            output = root / "out"
            prepare_manual_review(source, root / "raw", output, expected_rows=1)
            review = output / "response_function_manual_review.csv"
            with review.open(newline="", encoding="utf-8") as handle:
                reviewed = list(csv.DictReader(handle))
            reviewed[0]["source_row_sha256"] = "0" * 64
            for column in LABEL_COLUMNS:
                reviewed[0][column] = "1"
            write_csv(review, reviewed)
            audit = validate_human_review(review, source, output, expected_rows=1)
            self.assertEqual(audit["status"], "FAIL_SOURCE_BINDING")
            self.assertEqual(
                audit["source_hash_mismatch_review_ids"],
                [reviewed[0]["review_id"]],
            )


if __name__ == "__main__":
    unittest.main()
