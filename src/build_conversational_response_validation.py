#!/usr/bin/env python3
"""Classify conversational-context mismatches and prepare manual labels.

This module never infers the registered manual outcomes. It creates stable,
hash-bound review rows and validates labels entered by a human reviewer.
"""

from __future__ import annotations

import argparse
import bisect
import csv
import gzip
import hashlib
import json
import os
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, TextIO

try:
    from src.build_conversational_eligibility_sample import (
        CARETAKER_SPEAKERS,
        CARETAKER_SPEAKERS_BY_DATASET,
        MainTier,
        _normalized,
        parse_chat_main_tiers,
    )
except ImportError:  # pragma: no cover - direct-script execution
    from build_conversational_eligibility_sample import (
        CARETAKER_SPEAKERS,
        CARETAKER_SPEAKERS_BY_DATASET,
        MainTier,
        _normalized,
        parse_chat_main_tiers,
    )


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_LINK_ROOT = ROOT / "results/external/portable_t7"
DEFAULT_FLAGS = DEFAULT_LINK_ROOT / "conversational_eligibility/full79_child_conversational_flags.csv.gz"
DEFAULT_SAMPLE = DEFAULT_LINK_ROOT / "conversational_eligibility/full79_conversational_manual_validation_sample.csv"
DEFAULT_RAW_ROOT = DEFAULT_LINK_ROOT / "raw_chat"
DEFAULT_OUTPUT = ROOT / "results/conversational_response_validation_20260904"

LABEL_COLUMNS = (
    "manual_genuine_response",
    "manual_imitation",
    "manual_routine_or_reading",
    "manual_backchannel_or_acknowledgement",
    "manual_repair_or_clarification",
    "manual_next_response_contingent",
)
ALLOWED_LABELS = {"0", "1", "U", "NA"}
IDENTITY_COLUMNS = ("dataset", "child_id", "file", "line_no", "utt_id")
EXPECTED_EMPTY_TIER_CLASSIFICATION = "expected_empty_immediate_caregiver_skipped"


def open_text(path: Path, mode: str = "rt") -> TextIO:
    if path.suffix == ".gz":
        return gzip.open(path, mode, encoding="utf-8", newline="")
    return path.open(mode, encoding="utf-8", newline="")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def atomic_json(payload: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    try:
        temporary.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def source_row_hash(row: dict[str, str]) -> str:
    payload = {key: value for key, value in sorted(row.items()) if not key.startswith("manual_")}
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def review_id(row: dict[str, str]) -> str:
    identity = "|".join(row.get(column, "") for column in IDENTITY_COLUMNS)
    return "RV-" + hashlib.sha256(identity.encode("utf-8")).hexdigest()[:16]


def _prior_match(
    tiers: list[MainTier],
    normalized_positions: dict[str, list[int]],
    target_index: int,
    context: str,
    caretakers: set[str],
) -> tuple[MainTier | None, int | None, int | None]:
    normalized_context = _normalized(context)
    positions = normalized_positions.get(normalized_context, []) if normalized_context else []
    offset = bisect.bisect_left(positions, target_index) - 1
    if offset < 0:
        return None, None, None
    index = positions[offset]
    tier = tiers[index]
    caregiver_rank = None
    if tier.speaker in caretakers:
        caregiver_rank = sum(
            candidate.speaker in caretakers for candidate in tiers[index : target_index]
        )
    return tier, target_index - index - 1, caregiver_rank


def classify_mismatch(
    row: dict[str, str],
    *,
    raw_root: Path,
    tier_cache: dict[Path, tuple[list[MainTier], dict[int, int], dict[str, list[int]]]],
) -> dict[str, str]:
    dataset = row.get("dataset", "")
    raw_path = raw_root / dataset / row.get("file", "")
    result = {
        "classification": "",
        "raw_source_current": str(raw_path),
        "immediate_previous_raw": "",
        "immediate_previous_clean_nonempty": "",
        "nearest_nonempty_caregiver_speaker": "",
        "nearest_nonempty_caregiver_line_no": "",
        "nearest_nonempty_caregiver_main_gap": "",
        "context_k1_matches_nearest_nonempty_caregiver": "",
        "matched_prior_speaker": "",
        "matched_prior_line_no": "",
        "matched_prior_main_gap": "",
        "matched_prior_caregiver_rank": "",
        "normalized_similarity": "",
    }
    if not raw_path.is_file():
        result["classification"] = "raw_file_missing"
        return result
    if raw_path not in tier_cache:
        tiers = parse_chat_main_tiers(raw_path)[0]
        line_indexes = {tier.line_no: index for index, tier in enumerate(tiers)}
        normalized_positions: dict[str, list[int]] = defaultdict(list)
        for index, tier in enumerate(tiers):
            normalized_positions[_normalized(tier.utterance_clean)].append(index)
        tier_cache[raw_path] = (tiers, line_indexes, dict(normalized_positions))
    tiers, indexes, normalized_positions = tier_cache[raw_path]
    try:
        line_no = int(float(row.get("line_no", "")))
    except ValueError:
        result["classification"] = "invalid_target_line"
        return result
    target_index = indexes.get(line_no)
    if target_index is None:
        result["classification"] = "target_line_missing"
        return result

    caretakers = CARETAKER_SPEAKERS_BY_DATASET.get(dataset, CARETAKER_SPEAKERS)
    previous = tiers[target_index - 1] if target_index else None
    context = row.get("context_k1", "")
    expected = previous.utterance_clean if previous else ""
    context_norm = _normalized(context)
    expected_norm = _normalized(expected)
    similarity = SequenceMatcher(None, context_norm, expected_norm).ratio() if context_norm or expected_norm else 1.0
    result["normalized_similarity"] = f"{similarity:.6f}"
    result["immediate_previous_raw"] = previous.utterance_raw if previous else ""
    result["immediate_previous_clean_nonempty"] = "1" if expected_norm else "0"

    if previous is None or previous.speaker not in caretakers:
        result["classification"] = "immediate_tier_not_allowed_caregiver"
        return result
    if context_norm == expected_norm:
        result["classification"] = "normalized_match_inconsistent_flag"
        return result

    nearest_nonempty_caregiver: MainTier | None = None
    nearest_nonempty_caregiver_gap: int | None = None
    for prior_index in range(target_index - 1, -1, -1):
        candidate = tiers[prior_index]
        if candidate.speaker in caretakers and _normalized(candidate.utterance_clean):
            nearest_nonempty_caregiver = candidate
            nearest_nonempty_caregiver_gap = target_index - prior_index - 1
            break
    context_matches_nearest_nonempty = bool(
        nearest_nonempty_caregiver is not None
        and context_norm == _normalized(nearest_nonempty_caregiver.utterance_clean)
    )
    result.update(
        {
            "nearest_nonempty_caregiver_speaker": (
                nearest_nonempty_caregiver.speaker if nearest_nonempty_caregiver else ""
            ),
            "nearest_nonempty_caregiver_line_no": (
                str(nearest_nonempty_caregiver.line_no) if nearest_nonempty_caregiver else ""
            ),
            "nearest_nonempty_caregiver_main_gap": (
                str(nearest_nonempty_caregiver_gap)
                if nearest_nonempty_caregiver_gap is not None
                else ""
            ),
            "context_k1_matches_nearest_nonempty_caregiver": (
                "1" if context_matches_nearest_nonempty else "0"
            ),
        }
    )
    if not expected_norm and context_matches_nearest_nonempty:
        result["classification"] = EXPECTED_EMPTY_TIER_CLASSIFICATION
        return result

    matched, main_gap, caregiver_rank = _prior_match(
        tiers, normalized_positions, target_index, context, caretakers
    )
    if matched is not None:
        result.update(
            {
                "matched_prior_speaker": matched.speaker,
                "matched_prior_line_no": str(matched.line_no),
                "matched_prior_main_gap": str(main_gap),
                "matched_prior_caregiver_rank": str(caregiver_rank or ""),
            }
        )
        if matched.speaker in caretakers:
            result["classification"] = "context_points_to_earlier_allowed_caregiver"
        elif matched.speaker == "CHI":
            result["classification"] = "context_points_to_prior_child"
        else:
            result["classification"] = "context_points_to_prior_other_speaker"
    elif expected_norm and expected_norm in context_norm:
        result["classification"] = "context_contains_immediate_caregiver"
    elif context_norm and context_norm in expected_norm:
        result["classification"] = "immediate_caregiver_contains_context"
    elif similarity >= 0.8:
        result["classification"] = "high_similarity_cleaning_difference"
    else:
        result["classification"] = "no_prior_exact_main_tier_match"
    return result


def write_csv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, str]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}{path.suffix}")
    count = 0
    try:
        with open_text(temporary, "wt") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
                count += 1
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)
    return count


def audit_context_mismatches(flags_path: Path, raw_root: Path, output_dir: Path) -> dict[str, Any]:
    detail_path = output_dir / "context_k1_mismatch_rows.csv.gz"
    summary_path = output_dir / "context_k1_mismatch_summary.csv"
    tier_cache: dict[
        Path, tuple[list[MainTier], dict[int, int], dict[str, list[int]]]
    ] = {}
    counts: Counter[str] = Counter()
    by_dataset: dict[str, Counter[str]] = defaultdict(Counter)
    detail_rows: list[dict[str, str]] = []
    with open_text(flags_path) as handle:
        reader = csv.DictReader(handle)
        input_columns = list(reader.fieldnames or [])
        for row in reader:
            counts["rows"] += 1
            if row.get("primary_responsive_turn_eligible") != "1":
                continue
            counts["eligible_rows"] += 1
            if row.get("context_k1_matches_nearest_caretaker") == "1":
                continue
            classified = classify_mismatch(row, raw_root=raw_root, tier_cache=tier_cache)
            label = classified["classification"]
            counts["mismatch_rows"] += 1
            counts[label] += 1
            if label == EXPECTED_EMPTY_TIER_CLASSIFICATION:
                counts["adjudicated_expected_rows"] += 1
            else:
                counts["unexpected_mismatch_rows"] += 1
            by_dataset[row.get("dataset", "")][label] += 1
            detail_rows.append(
                {
                    "review_id": review_id(row),
                    **{column: row.get(column, "") for column in input_columns},
                    **classified,
                }
            )
    extra_columns = [
        "classification",
        "raw_source_current",
        "immediate_previous_raw",
        "immediate_previous_clean_nonempty",
        "nearest_nonempty_caregiver_speaker",
        "nearest_nonempty_caregiver_line_no",
        "nearest_nonempty_caregiver_main_gap",
        "context_k1_matches_nearest_nonempty_caregiver",
        "matched_prior_speaker",
        "matched_prior_line_no",
        "matched_prior_main_gap",
        "matched_prior_caregiver_rank",
        "normalized_similarity",
    ]
    write_csv(detail_path, ["review_id", *input_columns, *extra_columns], detail_rows)
    summary_rows = []
    for dataset, values in sorted(by_dataset.items()):
        for classification, count in sorted(values.items()):
            summary_rows.append(
                {"dataset": dataset, "classification": classification, "rows": str(count)}
            )
    write_csv(summary_path, ["dataset", "classification", "rows"], summary_rows)
    for key in (
        "mismatch_rows",
        "adjudicated_expected_rows",
        "unexpected_mismatch_rows",
    ):
        counts[key] += 0
    if counts["mismatch_rows"] == 0:
        status = "PASS_NO_MISMATCHES"
    elif counts["unexpected_mismatch_rows"] == 0:
        status = "PASS_ADJUDICATED_EXPECTED_EMPTY_TIERS"
    else:
        status = "REVIEW_UNEXPECTED_MISMATCHES"
    audit = {
        "status": status,
        "flags_path": str(flags_path),
        "flags_sha256": sha256_file(flags_path),
        "raw_root": str(raw_root),
        "counts": dict(sorted(counts.items())),
        "parsed_raw_files": len(tier_cache),
        "detail_path": str(detail_path),
        "detail_sha256": sha256_file(detail_path),
        "summary_path": str(summary_path),
        "summary_sha256": sha256_file(summary_path),
        "adjudication": {
            "expected_classification": EXPECTED_EMPTY_TIER_CLASSIFICATION,
            "interpretation": (
                "The immediately preceding allowed-caregiver main tier cleans to empty, so "
                "the scorer correctly uses the nearest prior non-empty allowed-caregiver tier."
            ),
            "strict_analysis_impact": (
                "These rows fail the existing context-match gate and are excluded from the "
                "strict caregiver-child-caregiver analysis."
            ),
        },
    }
    atomic_json(audit, output_dir / "context_k1_mismatch_audit.json")
    return audit


def manual_label_audit(rows: list[dict[str, str]], expected_rows: int = 325) -> dict[str, Any]:
    identifiers = [row.get("review_id", "") or review_id(row) for row in rows]
    invalid: list[dict[str, str]] = []
    rows_any = 0
    rows_complete = 0
    for row_id, row in zip(identifiers, rows):
        values = [row.get(column, "").strip().upper() for column in LABEL_COLUMNS]
        rows_any += int(any(values))
        rows_complete += int(all(value in ALLOWED_LABELS for value in values))
        for column, value in zip(LABEL_COLUMNS, values):
            if value and value not in ALLOWED_LABELS:
                invalid.append({"review_id": row_id, "column": column, "value": value})
    duplicates = len(identifiers) - len(set(identifiers))
    ready = len(rows) == expected_rows and rows_complete == expected_rows and not invalid and duplicates == 0
    if ready:
        status = "PASS_HUMAN_LABELS_COMPLETE"
    elif rows_any:
        status = "IN_PROGRESS"
    else:
        status = "READY_FOR_HUMAN_REVIEW"
    return {
        "status": status,
        "expected_rows": expected_rows,
        "rows": len(rows),
        "rows_with_any_label": rows_any,
        "rows_with_all_required_labels": rows_complete,
        "duplicate_review_ids": duplicates,
        "invalid_labels": invalid,
        "allowed_labels": sorted(ALLOWED_LABELS),
        "required_label_columns": list(LABEL_COLUMNS),
    }


def read_review_rows(path: Path) -> list[dict[str, str]]:
    """Read a human-edited review CSV or the Review sheet of its workbook."""
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook["Review"]
            values = worksheet.iter_rows(values_only=True)
            header = [str(value or "") for value in next(values)]
            return [
                {
                    column: "" if value is None else str(value)
                    for column, value in zip(header, row)
                }
                for row in values
                if any(value is not None and str(value) != "" for value in row)
            ]
        finally:
            workbook.close()
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def validate_human_review(
    review_path: Path,
    sample_path: Path,
    output_dir: Path,
    expected_rows: int = 325,
) -> dict[str, Any]:
    """Validate labels and prove that every review row belongs to the sample."""
    with sample_path.open(newline="", encoding="utf-8") as handle:
        source_rows = list(csv.DictReader(handle))
    expected = {review_id(row): source_row_hash(row) for row in source_rows}
    rows = read_review_rows(review_path)
    audit = manual_label_audit(rows, expected_rows=expected_rows)
    observed_ids = {row.get("review_id", "") for row in rows}
    missing_ids = sorted(set(expected) - observed_ids)
    unknown_ids = sorted(observed_ids - set(expected))
    hash_mismatches = sorted(
        row_id
        for row in rows
        if (row_id := row.get("review_id", "")) in expected
        and row.get("source_row_sha256", "") != expected[row_id]
    )
    binding_passed = (
        len(source_rows) == expected_rows
        and not missing_ids
        and not unknown_ids
        and not hash_mismatches
        and audit["duplicate_review_ids"] == 0
    )
    if not binding_passed:
        audit["status"] = "FAIL_SOURCE_BINDING"
    audit.update(
        {
            "source_binding_passed": binding_passed,
            "missing_review_ids": missing_ids,
            "unknown_review_ids": unknown_ids,
            "source_hash_mismatch_review_ids": hash_mismatches,
            "source_sample": str(sample_path),
            "source_sample_sha256": sha256_file(sample_path),
            "review_file": str(review_path),
            "review_file_sha256": sha256_file(review_path),
        }
    )
    atomic_json(audit, output_dir / "manual_label_audit.json")
    return audit


def render_codebook(path: Path) -> None:
    path.write_text(
        """# Response-function manual review codebook

Use `1` for yes, `0` for no, `U` when the transcript is genuinely ambiguous,
and `NA` only when the outcome cannot apply (for example, no immediate next
caregiver exists for the next-response label). Every required label cell must
be completed by a human reviewer before the response-function model gate can
pass.

After saving the workbook, validate it without regenerating or overwriting it:

```bash
.venv/bin/python src/build_conversational_response_validation.py \\
  --stage validate-labels \\
  --review-file results/conversational_response_validation_20260904/response_function_manual_review.xlsx
```

- `manual_genuine_response`: the child turn is meaningfully responsive to the
  immediately preceding main tier.
- `manual_imitation`: the child substantially repeats the preceding speaker,
  rather than merely sharing a function word.
- `manual_routine_or_reading`: the exchange is embedded in a scripted routine,
  song, recitation, or book-reading sequence.
- `manual_backchannel_or_acknowledgement`: the child primarily acknowledges or
  minimally signals attention/acceptance.
- `manual_repair_or_clarification`: the child or immediately adjacent caregiver
  turn performs repair or requests/provides clarification.
- `manual_next_response_contingent`: the immediately following caregiver turn
  is semantically contingent on the child's utterance.
- `manual_notes`: briefly justify `U`, unusual `NA`, or difficult decisions.

Candidate columns are screening aids only. Do not copy them mechanically into
manual labels. Review the raw child turn and both adjacent main tiers.
""",
        encoding="utf-8",
    )


def write_workbook(rows: list[dict[str, str]], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["Response-function manual validation"])
    instructions.append(["Allowed labels", "1 = yes; 0 = no; U = uncertain; NA = not applicable"])
    instructions.append(["Rule", "Do not copy candidate flags mechanically; inspect the adjacent turns."])
    instructions["A1"].font = Font(bold=True, size=14)
    data = workbook.create_sheet("Review")
    columns = list(rows[0]) if rows else ["review_id", *LABEL_COLUMNS]
    data.append(columns)
    for row in rows:
        data.append(
            [
                ILLEGAL_CHARACTERS_RE.sub("", str(row.get(column, "")))
                for column in columns
            ]
        )
    data.freeze_panes = "A2"
    data.auto_filter.ref = data.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    label_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in data[1]:
        cell.font = Font(bold=True)
        cell.fill = label_fill if cell.value in LABEL_COLUMNS else header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    validation = DataValidation(type="list", formula1='"0,1,U,NA"', allow_blank=True)
    data.add_data_validation(validation)
    for column_number, column in enumerate(columns, start=1):
        letter = data.cell(row=1, column=column_number).column_letter
        if column in LABEL_COLUMNS:
            validation.add(f"{letter}2:{letter}{len(rows) + 1}")
            data.column_dimensions[letter].width = 18
        elif column in {
            "child_utterance_raw",
            "previous_main_utterance_clean",
            "next_main_utterance_clean",
            "context_k1",
            "manual_notes",
        }:
            data.column_dimensions[letter].width = 42
        else:
            data.column_dimensions[letter].width = min(max(len(column) + 2, 10), 24)
    for row in data.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def prepare_manual_review(
    sample_path: Path, raw_root: Path, output_dir: Path, expected_rows: int = 325
) -> dict[str, Any]:
    with sample_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        source_columns = list(reader.fieldnames or [])
        source_rows = list(reader)
    rows = []
    for source in source_rows:
        row = {
            "review_id": review_id(source),
            "source_row_sha256": source_row_hash(source),
            "raw_source_current": str(raw_root / source.get("dataset", "") / source.get("file", "")),
            **source,
        }
        rows.append(row)
    output_csv = output_dir / "response_function_manual_review.csv"
    output_xlsx = output_dir / "response_function_manual_review.xlsx"
    codebook = output_dir / "response_function_manual_review_codebook.md"
    write_csv(output_csv, ["review_id", "source_row_sha256", "raw_source_current", *source_columns], rows)
    write_workbook(rows, output_xlsx)
    render_codebook(codebook)
    audit = manual_label_audit(rows, expected_rows=expected_rows)
    audit.update(
        {
            "source_sample": str(sample_path),
            "source_sample_sha256": sha256_file(sample_path),
            "review_csv": str(output_csv),
            "review_csv_sha256": sha256_file(output_csv),
            "review_workbook": str(output_xlsx),
            "review_workbook_sha256": sha256_file(output_xlsx),
            "codebook": str(codebook),
            "codebook_sha256": sha256_file(codebook),
        }
    )
    atomic_json(audit, output_dir / "manual_label_audit.json")
    return audit


def render_report(mismatch: dict[str, Any], manual: dict[str, Any], path: Path) -> None:
    counts = mismatch["counts"]
    lines = [
        "# Conversational response validation status",
        "",
        f"- Context audit: `{mismatch['status']}`",
        f"- Structurally eligible rows: `{int(counts.get('eligible_rows', 0)):,}`",
        f"- Classified context-k1 mismatches: `{int(counts.get('mismatch_rows', 0)):,}`",
        f"- Adjudicated empty-tier skips: `{int(counts.get('adjudicated_expected_rows', 0)):,}`",
        f"- Unexpected context mismatches: `{int(counts.get('unexpected_mismatch_rows', 0)):,}`",
        f"- Manual-label gate: `{manual['status']}`",
        f"- Review rows: `{manual['rows']:,}`",
        f"- Fully labeled rows: `{manual['rows_with_all_required_labels']:,}`",
        "",
        "The mismatch table is an adjudication aid, not a new scientific outcome.",
        "Response-function models remain blocked until all required fields are human-reviewed.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--stage",
        choices=("all", "mismatches", "manual", "validate-labels"),
        default="all",
    )
    parser.add_argument("--flags", type=Path, default=DEFAULT_FLAGS)
    parser.add_argument("--manual-sample", type=Path, default=DEFAULT_SAMPLE)
    parser.add_argument("--raw-root", type=Path, default=DEFAULT_RAW_ROOT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--review-file",
        type=Path,
        help="human-edited CSV/XLSX to check during --stage validate-labels",
    )
    parser.add_argument("--expected-manual-rows", type=int, default=325)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    mismatch_path = args.output_dir / "context_k1_mismatch_audit.json"
    manual_path = args.output_dir / "manual_label_audit.json"
    if args.stage in {"all", "mismatches"}:
        mismatch = audit_context_mismatches(args.flags, args.raw_root, args.output_dir)
    else:
        mismatch = json.loads(mismatch_path.read_text(encoding="utf-8"))
    if args.stage in {"all", "manual"}:
        manual = prepare_manual_review(
            args.manual_sample,
            args.raw_root,
            args.output_dir,
            expected_rows=args.expected_manual_rows,
        )
    elif args.stage == "validate-labels":
        review_path = args.review_file or (
            args.output_dir / "response_function_manual_review.xlsx"
        )
        manual = validate_human_review(
            review_path,
            args.manual_sample,
            args.output_dir,
            expected_rows=args.expected_manual_rows,
        )
    elif manual_path.is_file():
        manual = json.loads(manual_path.read_text(encoding="utf-8"))
    else:
        manual = {
            "status": "NOT_RUN",
            "rows": 0,
            "rows_with_all_required_labels": 0,
        }
    render_report(mismatch, manual, args.output_dir / "README.md")
    print(json.dumps({"mismatch": mismatch, "manual": manual}, indent=2, sort_keys=True))
    mismatch_passed = mismatch["status"].startswith("PASS_")
    if args.stage == "validate-labels":
        labels_passed = manual["status"] == "PASS_HUMAN_LABELS_COMPLETE"
        return 0 if mismatch_passed and labels_passed else 1
    return 0 if mismatch_passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
