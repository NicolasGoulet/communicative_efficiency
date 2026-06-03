#!/usr/bin/env python3
"""Build a publication-oriented utterance measurement validation probe.

The output is intentionally redundant: it reports several strategies side by
side so that word, morpheme, syllable, and phoneme counts can be audited before
any Route 1 model uses them.

Reference hierarchy used here:

1. Words: cleaned orthographic lexical tokens, excluding punctuation.
2. Morphemes: CHAT `%mor` tier when available; suffix heuristics as fallback.
3. Syllables: CMU Pronouncing Dictionary ARPABET vowel/stress phones when a
   word is in-dictionary; the `syllables` package for OOV words as written.
4. Phonemes: CMUdict ARPABET segment count when a word is in-dictionary;
   g2p-en predictions for OOV words as written.

Rows with OOV words are flagged, and the recommended columns record the fallback
used for those word forms.
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import re
import warnings
from dataclasses import asdict, dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
import pronouncing
import pyphen
import syllables

from utterance_count_strategies import (
    count_morphemes_suffix_heuristic,
    count_syllables,
    count_words_regex,
    count_words_whitespace,
    normalize_text,
    word_syllables_final_le,
    word_tokens_regex,
)


DEFAULT_DATASETS = ("Brown", "Manchester", "Providence")
DEFAULT_SCORING_ROOT = Path("data/big_cleaned_dataset/default_naturalistic_merged_006_023/preprocessed_data")
DEFAULT_RAW_ROOT = Path("data/raw_data")
DEFAULT_OUTPUT_CSV = Path("results/count_validation/publication_measurement_probe_50.csv")
DEFAULT_OUTPUT_MD = Path("results/count_validation/publication_measurement_probe_50.md")
DEFAULT_REVIEW_CSV = Path("results/count_validation/publication_measurement_review_50.csv")
DEFAULT_TOKEN_CSV = Path("results/count_validation/publication_measurement_review_50_tokens.csv")
DEFAULT_REVIEW_XLSX = Path("results/count_validation/publication_measurement_review_50.xlsx")
PROJECT_ROOT = Path(__file__).resolve().parents[1]
LOCAL_NLTK_DATA = PROJECT_ROOT / "data" / "nltk_data"

PUNCT_MOR_RE = re.compile(r"^[.?!,;:]+$")
CHAT_TIME_RE = re.compile(r"\x15.*?\x15")
MOR_TOKEN_SPLIT_RE = re.compile(r"\s+")
CONTRACTION_RE = re.compile(r"\b\w+'(?:m|re|ve|ll|d|s|t)\b|n't\b", re.IGNORECASE)
BOUND_SUFFIX_RE = re.compile(r"\b\w+(?:ing|ed|es|s)\b", re.IGNORECASE)
CHAT_REPETITION_MARKER_RE = re.compile(r"\[(?:/|//|x\s*\d+)\]")
ARPABET_PHONE_RE = re.compile(r"^[A-Z]+[0-2]?$")
STRESS_PHONE_RE = re.compile(r"\d$")

pyphen_dic = pyphen.Pyphen(lang="en_US")


@dataclass(frozen=True)
class CmuWordPronunciation:
    word: str
    phones_first: str
    pronunciation_count: int
    syllables_first: int | None
    syllables_min: int | None
    syllables_max: int | None
    phonemes_first: int | None
    phonemes_min: int | None
    phonemes_max: int | None


@dataclass(frozen=True)
class CmuUtteranceCounts:
    cmu_syllable_count: int | None
    cmu_syllable_min: int | None
    cmu_syllable_max: int | None
    cmu_phoneme_count: int | None
    cmu_phoneme_min: int | None
    cmu_phoneme_max: int | None
    cmu_all_words_in_dict: int
    cmu_oov_count: int
    cmu_oov_words: str
    cmu_ambiguous_pronunciation_words: str
    cmu_word_pronunciations_json: str


@dataclass(frozen=True)
class HybridWordPronunciation:
    word: str
    source: str
    phones: str
    syllables: int
    phonemes: int


@dataclass(frozen=True)
class HybridUtteranceCounts:
    hybrid_syllable_count: int
    hybrid_phoneme_count: int
    hybrid_g2p_fallback_word_count: int
    hybrid_g2p_fallback_words: str
    hybrid_word_pronunciations_json: str


@dataclass(frozen=True)
class SyllableWordCount:
    word: str
    source: str
    syllables: int


@dataclass(frozen=True)
class SyllableUtteranceCounts:
    syllable_count: int
    fallback_word_count: int
    fallback_words: str
    word_counts_json: str


@dataclass(frozen=True)
class MorCounts:
    mor_tier_found: int
    mor_tier: str
    mor_component_count: int | None
    mor_mlu_proxy_count: int | None
    mor_bound_morpheme_tags: str


@dataclass(frozen=True)
class ProbeRow:
    dataset: str
    child_id: str
    session_id: str
    age_months: str
    file: str
    line_no: str
    utt_id: str
    raw_main_tier: str
    utterance_clean: str
    word_tokens_json: str
    word_count_regex: int
    word_count_whitespace: int
    mor_tier_found: int
    mor_component_count: int | None
    mor_mlu_proxy_count: int | None
    morpheme_count_suffix_heuristic: int
    mor_bound_morpheme_tags: str
    syllable_count_cmudict: int | None
    syllable_count_cmudict_min: int | None
    syllable_count_cmudict_max: int | None
    syllable_count_pyphen: int
    syllable_count_syllables_pkg: int
    syllable_count_vowel_le_heuristic: int
    phoneme_count_cmudict: int | None
    phoneme_count_cmudict_min: int | None
    phoneme_count_cmudict_max: int | None
    syllable_count_cmudict_or_syllables_pkg: int
    syllable_pkg_fallback_word_count: int
    syllable_pkg_fallback_words: str
    syllable_word_counts_json: str
    syllable_count_cmudict_g2p: int
    phoneme_count_cmudict_g2p: int
    g2p_fallback_word_count: int
    g2p_fallback_words: str
    hybrid_word_pronunciations_json: str
    cmu_all_words_in_dict: int
    cmu_oov_count: int
    cmu_oov_words: str
    cmu_ambiguous_pronunciation_words: str
    cmu_word_pronunciations_json: str
    count_disagreement_score: int
    raw_repetition_marker: int
    mor_surface_mismatch: int
    recommended_word_count: int
    recommended_morpheme_count: int | None
    recommended_syllable_count: int | None
    recommended_phoneme_count: int | None
    quality_flags: str
    mor_tier: str


def clean_raw_main_tier(line: str) -> str:
    """Remove CHAT time codes and normalize a raw main-tier line."""
    if not line:
        return ""
    return normalize_text(CHAT_TIME_RE.sub(" ", line))


def raw_chat_path(raw_root: Path, dataset: str, file_value: str) -> Path:
    """Resolve a raw CHAT file from scoring CSV provenance."""
    return raw_root / dataset / file_value


@lru_cache(maxsize=2048)
def read_raw_chat_lines(path_value: str) -> tuple[str, ...]:
    """Read a CHAT file once and reuse it across many row lookups."""
    path = Path(path_value)
    if not path.is_file():
        return ()
    return tuple(path.read_text(encoding="utf-8", errors="replace").splitlines())


def read_raw_record(raw_root: Path, dataset: str, file_value: str, line_no: str) -> tuple[str, str]:
    """Return the raw main tier and following `%mor` tier for a scored row."""
    path = raw_chat_path(raw_root, dataset, file_value)
    try:
        target_idx = int(float(str(line_no))) - 1
    except ValueError:
        return "", ""
    lines = read_raw_chat_lines(str(path))
    if target_idx < 0 or target_idx >= len(lines):
        return "", ""

    main = clean_raw_main_tier(lines[target_idx])
    mor_parts: list[str] = []
    for line in lines[target_idx + 1 :]:
        if line.startswith("*") or line.startswith("@"):
            break
        if line.startswith("%mor:"):
            mor_parts.append(line.split(":", 1)[1].strip())
            continue
        if mor_parts and (line.startswith("\t") or line.startswith(" ")):
            mor_parts.append(line.strip())
            continue
        if mor_parts:
            break
    return main, normalize_text(" ".join(mor_parts))


def mor_components(mor_tier: str) -> list[str]:
    """Return countable MOR components, splitting contractions at `~`."""
    components: list[str] = []
    for token in MOR_TOKEN_SPLIT_RE.split(normalize_text(mor_tier)):
        if not token or PUNCT_MOR_RE.fullmatch(token) or token in {"+...", "+..", "+/."}:
            continue
        if token == "cm|cm":
            continue
        for part in token.split("~"):
            part = part.strip()
            if not part or PUNCT_MOR_RE.fullmatch(part) or part == "cm|cm":
                continue
            if "|" not in part:
                continue
            components.append(part)
    return components


def mor_component_extra_bound_tags(component: str) -> list[str]:
    """
    Approximate overt English bound morphemes from a MOR component.

    This is a transparent proxy for inspection. For final MLU-style analyses,
    CLAN's `mlu` command on `%mor` tiers remains the domain-native reference.
    """
    pos, rest = component.split("|", 1)
    bits = [bit for bit in rest.split("-") if bit]
    if len(bits) <= 1:
        return []
    lemma = bits[0].lower()
    tags = set(bits[1:])
    out: list[str] = []

    if "Plur" in tags:
        out.append("Plur")
    if "Ger" in tags:
        out.append("Ger")
    if "Past" in tags and "irr" not in tags:
        out.append("Past")
    if pos == "verb" and "Pres" in tags and "S3" in tags and "irr" not in tags:
        out.append("3S")
    if ("Cmp" in tags or "Sup" in tags) and "irr" not in tags:
        out.extend(tag for tag in ("Cmp", "Sup") if tag in tags)
    if "Poss" in tags and lemma not in {"my", "your", "his", "her", "our", "their"}:
        out.append("Poss")
    return out


def count_mor_tier(mor_tier: str) -> MorCounts:
    """Count MOR components and a transparent MLU-style proxy."""
    if not mor_tier:
        return MorCounts(0, "", None, None, "")
    components = mor_components(mor_tier)
    bound_tags: list[str] = []
    total = 0
    for component in components:
        total += 1
        extras = mor_component_extra_bound_tags(component)
        total += len(extras)
        bound_tags.extend(extras)
    return MorCounts(
        mor_tier_found=1,
        mor_tier=mor_tier,
        mor_component_count=len(components),
        mor_mlu_proxy_count=total,
        mor_bound_morpheme_tags=";".join(bound_tags),
    )


@lru_cache(maxsize=50000)
def first_cmu_pronunciation(word: str) -> CmuWordPronunciation:
    """Return CMUdict pronunciation counts for one token."""
    phones = pronouncing.phones_for_word(word.lower())
    if not phones:
        return CmuWordPronunciation(word, "", 0, None, None, None, None, None, None)
    syll_counts = [pronouncing.syllable_count(phone) for phone in phones]
    phoneme_counts = [len(phone.split()) for phone in phones]
    return CmuWordPronunciation(
        word=word,
        phones_first=phones[0],
        pronunciation_count=len(phones),
        syllables_first=syll_counts[0],
        syllables_min=min(syll_counts),
        syllables_max=max(syll_counts),
        phonemes_first=phoneme_counts[0],
        phonemes_min=min(phoneme_counts),
        phonemes_max=max(phoneme_counts),
    )


@lru_cache(maxsize=1)
def g2p_model():
    """Create the g2p-en model, preferring repo-local NLTK resources."""
    if LOCAL_NLTK_DATA.exists():
        import nltk

        local_path = str(LOCAL_NLTK_DATA)
        if local_path not in nltk.data.path:
            nltk.data.path.insert(0, local_path)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=ResourceWarning, module=r"g2p_en\.g2p")
        from g2p_en import G2p

        return G2p()


def phones_syllable_count(phones: Sequence[str]) -> int:
    """Count ARPABET vowel/stress phones."""
    return sum(1 for phone in phones if STRESS_PHONE_RE.search(phone))


def normalize_g2p_phones(raw_phones: Sequence[str]) -> list[str]:
    """Keep ARPABET phones from g2p-en output and drop spaces/punctuation."""
    return [phone for phone in raw_phones if ARPABET_PHONE_RE.fullmatch(phone)]


@lru_cache(maxsize=50000)
def g2p_pronunciation_for_word(word: str) -> tuple[str, ...]:
    """Predict an ARPABET pronunciation for an OOV word form as written."""
    return tuple(normalize_g2p_phones(g2p_model()(word)))


def count_cmudict_g2p_utterance(tokens: Sequence[str]) -> HybridUtteranceCounts:
    """Use CMUdict for known words and g2p-en for OOV word forms as written."""
    word_prons: list[HybridWordPronunciation] = []
    g2p_words: list[str] = []
    for token in tokens:
        cmu = first_cmu_pronunciation(token)
        if cmu.pronunciation_count:
            phones = cmu.phones_first.split()
            source = "cmudict"
        else:
            phones = list(g2p_pronunciation_for_word(token))
            source = "g2p_en"
            g2p_words.append(token)
        syllable_count = phones_syllable_count(phones)
        if syllable_count <= 0 and phones:
            syllable_count = 1
        word_prons.append(
            HybridWordPronunciation(
                word=token,
                source=source,
                phones=" ".join(phones),
                syllables=syllable_count,
                phonemes=len(phones),
            )
        )
    return HybridUtteranceCounts(
        hybrid_syllable_count=sum(pron.syllables for pron in word_prons),
        hybrid_phoneme_count=sum(pron.phonemes for pron in word_prons),
        hybrid_g2p_fallback_word_count=len(g2p_words),
        hybrid_g2p_fallback_words=";".join(g2p_words),
        hybrid_word_pronunciations_json=json.dumps([asdict(pron) for pron in word_prons], ensure_ascii=False),
    )


def count_cmudict_or_syllables_pkg_utterance(tokens: Sequence[str]) -> SyllableUtteranceCounts:
    """Use CMUdict syllables for known words and orthographic estimates for OOV forms."""
    word_counts: list[SyllableWordCount] = []
    fallback_words: list[str] = []
    for token in tokens:
        cmu = first_cmu_pronunciation(token)
        if cmu.pronunciation_count and cmu.syllables_first is not None:
            source = "cmudict"
            syllable_count = int(cmu.syllables_first)
        else:
            source = "syllables_pkg"
            syllable_count = max(1, int(syllables.estimate(token.lower())))
            fallback_words.append(token)
        word_counts.append(SyllableWordCount(word=token, source=source, syllables=syllable_count))
    return SyllableUtteranceCounts(
        syllable_count=sum(count.syllables for count in word_counts),
        fallback_word_count=len(fallback_words),
        fallback_words=";".join(fallback_words),
        word_counts_json=json.dumps([asdict(count) for count in word_counts], ensure_ascii=False),
    )


def count_cmu_utterance(tokens: Sequence[str]) -> CmuUtteranceCounts:
    """Aggregate CMUdict syllable and phoneme counts for an utterance."""
    word_prons = [first_cmu_pronunciation(token) for token in tokens]
    oov = [pron.word for pron in word_prons if pron.pronunciation_count == 0]
    ambiguous = [pron.word for pron in word_prons if pron.pronunciation_count > 1]
    serializable = [asdict(pron) for pron in word_prons]
    if oov:
        return CmuUtteranceCounts(
            cmu_syllable_count=None,
            cmu_syllable_min=None,
            cmu_syllable_max=None,
            cmu_phoneme_count=None,
            cmu_phoneme_min=None,
            cmu_phoneme_max=None,
            cmu_all_words_in_dict=0,
            cmu_oov_count=len(oov),
            cmu_oov_words=";".join(oov),
            cmu_ambiguous_pronunciation_words=";".join(ambiguous),
            cmu_word_pronunciations_json=json.dumps(serializable, ensure_ascii=False),
        )
    syll_first = sum(int(pron.syllables_first or 0) for pron in word_prons)
    syll_min = sum(int(pron.syllables_min or 0) for pron in word_prons)
    syll_max = sum(int(pron.syllables_max or 0) for pron in word_prons)
    phon_first = sum(int(pron.phonemes_first or 0) for pron in word_prons)
    phon_min = sum(int(pron.phonemes_min or 0) for pron in word_prons)
    phon_max = sum(int(pron.phonemes_max or 0) for pron in word_prons)
    return CmuUtteranceCounts(
        cmu_syllable_count=syll_first,
        cmu_syllable_min=syll_min,
        cmu_syllable_max=syll_max,
        cmu_phoneme_count=phon_first,
        cmu_phoneme_min=phon_min,
        cmu_phoneme_max=phon_max,
        cmu_all_words_in_dict=1,
        cmu_oov_count=0,
        cmu_oov_words="",
        cmu_ambiguous_pronunciation_words=";".join(ambiguous),
        cmu_word_pronunciations_json=json.dumps(serializable, ensure_ascii=False),
    )


def count_pyphen_syllables(tokens: Sequence[str]) -> int:
    """Count orthographic hyphenation chunks with Pyphen."""
    total = 0
    for token in tokens:
        inserted = pyphen_dic.inserted(token.lower())
        total += max(1, len([part for part in inserted.split("-") if part]))
    return total


def count_syllables_pkg(tokens: Sequence[str]) -> int:
    """Count syllables with the `syllables` package heuristic."""
    return sum(max(1, int(syllables.estimate(token.lower()))) for token in tokens)


def nullable_int(value: int | None) -> int | None:
    return value if value is not None else None


def row_counts(
    row: pd.Series,
    *,
    raw_root: Path,
) -> ProbeRow:
    """Compute all measurement strategies for one scoring row."""
    utterance = normalize_text(row.get("chi_utterance_clean", row.get("utterance_clean", "")))
    tokens = word_tokens_regex(utterance)
    cmu = count_cmu_utterance(tokens)
    hybrid = count_cmudict_g2p_utterance(tokens)
    syllable_hybrid = count_cmudict_or_syllables_pkg_utterance(tokens)
    raw_main, mor_tier = read_raw_record(
        raw_root,
        str(row.get("dataset", "")),
        str(row.get("file", "")),
        str(row.get("line_no", "")),
    )
    mor = count_mor_tier(mor_tier)
    suffix_morph = count_morphemes_suffix_heuristic(utterance)
    vowel_le = count_syllables(utterance, word_syllables_final_le)
    pyphen_syll = count_pyphen_syllables(tokens)
    syll_pkg = count_syllables_pkg(tokens)

    quality_flags: list[str] = []
    if not mor.mor_tier_found:
        quality_flags.append("no_mor_tier")
    if not cmu.cmu_all_words_in_dict:
        quality_flags.append("cmu_oov")
    if hybrid.hybrid_g2p_fallback_word_count:
        quality_flags.append("g2p_fallback_used")
    if syllable_hybrid.fallback_word_count:
        quality_flags.append("syllable_pkg_fallback_used")
    if cmu.cmu_ambiguous_pronunciation_words:
        quality_flags.append("cmu_multiple_pronunciations")
    raw_repetition_marker = int(bool(CHAT_REPETITION_MARKER_RE.search(raw_main)))
    if raw_repetition_marker:
        quality_flags.append("raw_repetition_marker")
    word_count = count_words_regex(utterance)
    mor_surface_mismatch = int(
        mor.mor_component_count is not None
        and word_count > 0
        and abs(mor.mor_component_count - word_count) >= max(3, int(0.25 * word_count))
    )
    if mor_surface_mismatch:
        quality_flags.append("mor_surface_mismatch")

    syllable_values = [
        value
        for value in [cmu.cmu_syllable_count, hybrid.hybrid_syllable_count, pyphen_syll, syll_pkg, vowel_le]
        if value is not None
    ]
    morph_values = [
        value
        for value in [mor.mor_mlu_proxy_count, suffix_morph]
        if value is not None
    ]
    disagreement = 0
    disagreement += int(count_words_regex(utterance) != count_words_whitespace(utterance))
    if syllable_values:
        disagreement += max(syllable_values) - min(syllable_values)
    if morph_values:
        disagreement += max(morph_values) - min(morph_values)
    disagreement += 4 * cmu.cmu_oov_count
    disagreement += 2 * len([w for w in cmu.cmu_ambiguous_pronunciation_words.split(";") if w])
    disagreement += 2 if CONTRACTION_RE.search(utterance) else 0
    disagreement += 1 if BOUND_SUFFIX_RE.search(utterance) else 0
    disagreement += 4 if raw_repetition_marker else 0
    disagreement += 4 if mor_surface_mismatch else 0

    return ProbeRow(
        dataset=str(row.get("dataset", "")),
        child_id=str(row.get("child_id", "")),
        session_id=str(row.get("session_id", "")),
        age_months=str(row.get("age_months", "")),
        file=str(row.get("file", "")),
        line_no=str(row.get("line_no", "")),
        utt_id=str(row.get("utt_id", "")),
        raw_main_tier=raw_main,
        utterance_clean=utterance,
        word_tokens_json=json.dumps(tokens, ensure_ascii=False),
        word_count_regex=count_words_regex(utterance),
        word_count_whitespace=count_words_whitespace(utterance),
        mor_tier_found=mor.mor_tier_found,
        mor_component_count=mor.mor_component_count,
        mor_mlu_proxy_count=mor.mor_mlu_proxy_count,
        morpheme_count_suffix_heuristic=suffix_morph,
        mor_bound_morpheme_tags=mor.mor_bound_morpheme_tags,
        syllable_count_cmudict=nullable_int(cmu.cmu_syllable_count),
        syllable_count_cmudict_min=nullable_int(cmu.cmu_syllable_min),
        syllable_count_cmudict_max=nullable_int(cmu.cmu_syllable_max),
        syllable_count_pyphen=pyphen_syll,
        syllable_count_syllables_pkg=syll_pkg,
        syllable_count_vowel_le_heuristic=vowel_le,
        phoneme_count_cmudict=nullable_int(cmu.cmu_phoneme_count),
        phoneme_count_cmudict_min=nullable_int(cmu.cmu_phoneme_min),
        phoneme_count_cmudict_max=nullable_int(cmu.cmu_phoneme_max),
        syllable_count_cmudict_or_syllables_pkg=syllable_hybrid.syllable_count,
        syllable_pkg_fallback_word_count=syllable_hybrid.fallback_word_count,
        syllable_pkg_fallback_words=syllable_hybrid.fallback_words,
        syllable_word_counts_json=syllable_hybrid.word_counts_json,
        syllable_count_cmudict_g2p=hybrid.hybrid_syllable_count,
        phoneme_count_cmudict_g2p=hybrid.hybrid_phoneme_count,
        g2p_fallback_word_count=hybrid.hybrid_g2p_fallback_word_count,
        g2p_fallback_words=hybrid.hybrid_g2p_fallback_words,
        hybrid_word_pronunciations_json=hybrid.hybrid_word_pronunciations_json,
        cmu_all_words_in_dict=cmu.cmu_all_words_in_dict,
        cmu_oov_count=cmu.cmu_oov_count,
        cmu_oov_words=cmu.cmu_oov_words,
        cmu_ambiguous_pronunciation_words=cmu.cmu_ambiguous_pronunciation_words,
        cmu_word_pronunciations_json=cmu.cmu_word_pronunciations_json,
        count_disagreement_score=disagreement,
        raw_repetition_marker=raw_repetition_marker,
        mor_surface_mismatch=mor_surface_mismatch,
        recommended_word_count=word_count,
        recommended_morpheme_count=mor.mor_mlu_proxy_count if mor.mor_tier_found else suffix_morph,
        recommended_syllable_count=syllable_hybrid.syllable_count,
        recommended_phoneme_count=hybrid.hybrid_phoneme_count,
        quality_flags=";".join(quality_flags),
        mor_tier=mor.mor_tier,
    )


def iter_scoring_csvs(scoring_root: Path, datasets: Sequence[str]) -> Iterable[Path]:
    """Yield child scoring CSVs for requested datasets."""
    for dataset in datasets:
        root = scoring_root / dataset
        if not root.is_dir():
            continue
        yield from sorted(root.glob("*/chi.surprisal_scoring.csv"))


def load_candidate_probe_rows(
    *,
    scoring_root: Path,
    raw_root: Path,
    datasets: Sequence[str],
    min_words: int,
) -> list[ProbeRow]:
    """Load and measure all candidate child utterances."""
    measured: list[ProbeRow] = []
    for csv_path in iter_scoring_csvs(scoring_root, datasets):
        df = pd.read_csv(csv_path, dtype=str, keep_default_na=False, low_memory=False)
        for _, row in df.iterrows():
            utterance = normalize_text(row.get("chi_utterance_clean", ""))
            if count_words_regex(utterance) < min_words:
                continue
            measured.append(row_counts(row, raw_root=raw_root))
    return measured


def select_probe_rows(rows: Sequence[ProbeRow], sample_size: int, seed: int) -> list[ProbeRow]:
    """
    Select a deterministic validation sample.

    Rows are stratified by surface word count so the human review includes
    short, medium, long, and very long utterances, with hard cases prioritized
    inside each length bucket.
    """
    if len(rows) <= sample_size:
        return list(rows)

    rng = random.Random(seed)
    bucket_targets = {
        "short_1_3": max(1, round(sample_size * 0.28)),
        "medium_4_8": max(1, round(sample_size * 0.28)),
        "long_9_20": max(1, round(sample_size * 0.24)),
        "very_long_21_plus": max(1, sample_size - round(sample_size * 0.28) - round(sample_size * 0.28) - round(sample_size * 0.24)),
    }
    selected: list[ProbeRow] = []
    selected_keys: set[tuple[str, str, str, str, str]] = set()
    selected_diversity_keys: set[tuple[str, str, tuple[str, ...]]] = set()

    def add_row(row: ProbeRow) -> bool:
        key = (row.dataset, row.child_id, row.file, row.line_no, row.utt_id)
        diversity_key = probe_diversity_key(row)
        if key in selected_keys or diversity_key in selected_diversity_keys:
            return False
        selected.append(row)
        selected_keys.add(key)
        selected_diversity_keys.add(diversity_key)
        return True

    for bucket, target in bucket_targets.items():
        pool = [row for row in rows if length_bucket(row) == bucket]
        hard_pool = sorted(pool, key=sort_probe_hard_first)
        random_pool = list(pool)
        rng.shuffle(random_pool)
        hard_target = max(1, target // 2)
        for row in hard_pool:
            if len([r for r in selected if length_bucket(r) == bucket]) >= hard_target:
                break
            add_row(row)
        for row in random_pool:
            if len([r for r in selected if length_bucket(r) == bucket]) >= target:
                break
            add_row(row)

    sorted_rows = sorted(
        rows,
        key=sort_probe_hard_first,
    )
    for row in sorted_rows:
        if len(selected) >= sample_size:
            break
        add_row(row)

    remaining = [
        row
        for row in rows
        if (row.dataset, row.child_id, row.file, row.line_no, row.utt_id) not in selected_keys
    ]
    rng.shuffle(remaining)
    if len(selected) < sample_size:
        for row in remaining:
            if len(selected) >= sample_size:
                break
            selected.append(row)
    return sorted(
        selected,
        key=lambda row: (
            LENGTH_BUCKET_ORDER[length_bucket(row)],
            row.dataset,
            row.child_id,
            float(row.age_months or 0),
            row.file,
            int(float(row.line_no or 0)),
        ),
    )


def sort_probe_hard_first(row: ProbeRow) -> tuple[object, ...]:
    return (
        -row.count_disagreement_score,
        -row.g2p_fallback_word_count,
        -row.raw_repetition_marker,
        -row.mor_surface_mismatch,
        row.dataset,
        row.child_id,
        float(row.age_months) if row.age_months else -1.0,
        row.file,
        int(float(row.line_no)) if row.line_no else -1,
    )


def probe_diversity_key(row: ProbeRow) -> tuple[str, str, tuple[str, ...]]:
    """Collapse near-duplicate long repetition patterns in the validation sample."""
    tokens = tuple(json.loads(row.word_tokens_json))
    if len(tokens) > 12:
        tokens = tokens[:12]
    return ("global", "", tuple(token.lower() for token in tokens))


def length_bucket(row: ProbeRow) -> str:
    """Return a human-review length bucket based on surface word count."""
    words = row.recommended_word_count
    if words <= 3:
        return "short_1_3"
    if words <= 8:
        return "medium_4_8"
    if words <= 20:
        return "long_9_20"
    return "very_long_21_plus"


def indexed_tokens(row: ProbeRow) -> str:
    """Return a compact token index aid for manual counting."""
    tokens = json.loads(row.word_tokens_json)
    return " | ".join(f"{idx}:{token}" for idx, token in enumerate(tokens, start=1))


def issue_tags(row: ProbeRow) -> str:
    """Return compact issue tags for review triage."""
    tags: list[str] = []
    if row.g2p_fallback_word_count:
        tags.append("g2p")
    if row.raw_repetition_marker:
        tags.append("repetition")
    if row.mor_surface_mismatch:
        tags.append("mor_surface_mismatch")
    if CONTRACTION_RE.search(row.utterance_clean):
        tags.append("contraction")
    if "cmu_multiple_pronunciations" in row.quality_flags:
        tags.append("multiple_cmu_pron")
    if not tags:
        tags.append("ordinary")
    return ";".join(tags)


def g2p_pronunciation_audit(row: ProbeRow) -> str:
    """Return only the G2P fallback pronunciations in a compact readable form."""
    try:
        pronunciations = json.loads(row.hybrid_word_pronunciations_json)
    except json.JSONDecodeError:
        return ""
    parts: list[str] = []
    for pron in pronunciations:
        if pron.get("source") != "g2p_en":
            continue
        parts.append(
            f"{pron.get('word')}={pron.get('phones')} "
            f"[syl={pron.get('syllables')}, phon={pron.get('phonemes')}]"
        )
    return " | ".join(parts)


REVIEW_COLUMNS = [
    "review_id",
    "length_bucket",
    "issue_tags",
    "utterance_clean",
    "indexed_tokens",
    "auto_words",
    "manual_words",
    "auto_morphemes_surface",
    "manual_morphemes_surface",
    "auto_morphemes_mor_proxy",
    "manual_morphemes_mor_proxy",
    "auto_syllables_cmu_or_pkg",
    "auto_syllables_g2p_vowels",
    "auto_syllables_pkg",
    "auto_syllables_vowel_heuristic",
    "auto_syllables_pyphen",
    "manual_syllables",
    "auto_phonemes_cmu_or_g2p",
    "manual_phonemes",
    "syllable_pkg_fallback_words",
    "g2p_fallback_words",
    "g2p_pronunciation_audit",
    "raw_repetition_marker",
    "mor_surface_mismatch",
    "quality_flags",
    "dataset",
    "child_id",
    "age_months",
    "file",
    "line_no",
    "raw_main_tier",
    "mor_tier",
]


TOKEN_COLUMNS = [
    "review_id",
    "token_index",
    "token",
    "pron_source",
    "phones",
    "syllables_cmu_or_pkg",
    "syllables_g2p_vowels",
    "syllables_pkg",
    "syllables_vowel_heuristic",
    "syllables_pyphen",
    "phonemes_cmu_or_g2p",
    "is_g2p_fallback",
    "is_syllable_pkg_fallback",
    "dataset",
    "child_id",
    "age_months",
    "file",
    "line_no",
]

LENGTH_BUCKET_ORDER = {
    "short_1_3": 0,
    "medium_4_8": 1,
    "long_9_20": 2,
    "very_long_21_plus": 3,
}


def review_row_dict(row: ProbeRow, review_id: int) -> dict[str, object]:
    """Return compact human-review columns with manual blanks beside autos."""
    return {
        "review_id": review_id,
        "length_bucket": length_bucket(row),
        "issue_tags": issue_tags(row),
        "utterance_clean": row.utterance_clean,
        "indexed_tokens": indexed_tokens(row),
        "auto_words": row.recommended_word_count,
        "manual_words": "",
        "auto_morphemes_surface": row.morpheme_count_suffix_heuristic,
        "manual_morphemes_surface": "",
        "auto_morphemes_mor_proxy": row.mor_mlu_proxy_count,
        "manual_morphemes_mor_proxy": "",
        "auto_syllables_cmu_or_pkg": row.recommended_syllable_count,
        "auto_syllables_g2p_vowels": row.syllable_count_cmudict_g2p,
        "auto_syllables_pkg": row.syllable_count_syllables_pkg,
        "auto_syllables_vowel_heuristic": row.syllable_count_vowel_le_heuristic,
        "auto_syllables_pyphen": row.syllable_count_pyphen,
        "manual_syllables": "",
        "auto_phonemes_cmu_or_g2p": row.recommended_phoneme_count,
        "manual_phonemes": "",
        "syllable_pkg_fallback_words": row.syllable_pkg_fallback_words,
        "g2p_fallback_words": row.g2p_fallback_words,
        "g2p_pronunciation_audit": g2p_pronunciation_audit(row),
        "raw_repetition_marker": row.raw_repetition_marker,
        "mor_surface_mismatch": row.mor_surface_mismatch,
        "quality_flags": row.quality_flags,
        "dataset": row.dataset,
        "child_id": row.child_id,
        "age_months": row.age_months,
        "file": row.file,
        "line_no": row.line_no,
        "raw_main_tier": row.raw_main_tier,
        "mor_tier": row.mor_tier,
    }


def token_row_dicts(row: ProbeRow, review_id: int) -> list[dict[str, object]]:
    """Return one row per token for long-utterance review."""
    try:
        pronunciations = json.loads(row.hybrid_word_pronunciations_json)
    except json.JSONDecodeError:
        pronunciations = []
    try:
        syllable_counts = json.loads(row.syllable_word_counts_json)
    except json.JSONDecodeError:
        syllable_counts = []
    out: list[dict[str, object]] = []
    for idx, pron in enumerate(pronunciations, start=1):
        source = str(pron.get("source", ""))
        token = str(pron.get("word", ""))
        syllable_source = ""
        syllables_cmu_or_pkg: object = ""
        if idx - 1 < len(syllable_counts):
            syllable_source = str(syllable_counts[idx - 1].get("source", ""))
            syllables_cmu_or_pkg = syllable_counts[idx - 1].get("syllables", "")
        out.append(
            {
                "review_id": review_id,
                "token_index": idx,
                "token": token,
                "pron_source": source,
                "phones": pron.get("phones", ""),
                "syllables_cmu_or_pkg": syllables_cmu_or_pkg,
                "syllables_g2p_vowels": pron.get("syllables", ""),
                "syllables_pkg": max(1, int(syllables.estimate(token.lower()))) if token else "",
                "syllables_vowel_heuristic": word_syllables_final_le(token) if token else "",
                "syllables_pyphen": count_pyphen_syllables([token]) if token else "",
                "phonemes_cmu_or_g2p": pron.get("phonemes", ""),
                "is_g2p_fallback": int(source == "g2p_en"),
                "is_syllable_pkg_fallback": int(syllable_source == "syllables_pkg"),
                "dataset": row.dataset,
                "child_id": row.child_id,
                "age_months": row.age_months,
                "file": row.file,
                "line_no": row.line_no,
            }
        )
    return out


def write_review_outputs(
    rows: Sequence[ProbeRow],
    *,
    review_csv: Path,
    token_csv: Path,
    review_xlsx: Path,
) -> None:
    """Write compact CSVs and a formatted spreadsheet for human review."""
    review_csv.parent.mkdir(parents=True, exist_ok=True)
    review_rows = [review_row_dict(row, idx) for idx, row in enumerate(rows, start=1)]
    token_rows = [
        token_row
        for idx, row in enumerate(rows, start=1)
        for token_row in token_row_dicts(row, idx)
    ]

    with review_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=REVIEW_COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(review_rows)

    with token_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=TOKEN_COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(token_rows)

    write_review_xlsx(review_rows, token_rows, rows, review_xlsx)


def write_review_xlsx(
    review_rows: Sequence[dict[str, object]],
    token_rows: Sequence[dict[str, object]],
    full_rows: Sequence[ProbeRow],
    output_xlsx: Path,
) -> None:
    """Write a wrapped, frozen-pane workbook for LibreOffice review."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "review_50"
    write_sheet(ws, REVIEW_COLUMNS, review_rows)
    manual_fill = PatternFill("solid", fgColor="FFF2CC")
    for col_idx, column in enumerate(REVIEW_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if column.startswith("manual_"):
            for cell in ws[letter]:
                cell.fill = manual_fill
        if column in {"utterance_clean", "indexed_tokens", "g2p_pronunciation_audit", "raw_main_tier", "mor_tier"}:
            ws.column_dimensions[letter].width = 42
        elif column in {
            "review_id",
            "auto_words",
            "manual_words",
            "auto_syllables_cmu_or_pkg",
            "auto_syllables_g2p_vowels",
            "auto_syllables_pkg",
            "auto_syllables_vowel_heuristic",
            "auto_syllables_pyphen",
            "manual_syllables",
            "auto_phonemes_cmu_or_g2p",
            "manual_phonemes",
        }:
            ws.column_dimensions[letter].width = 13
        else:
            ws.column_dimensions[letter].width = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    token_ws = wb.create_sheet("tokens")
    write_sheet(token_ws, TOKEN_COLUMNS, token_rows)
    token_ws.freeze_panes = "A2"
    token_ws.auto_filter.ref = token_ws.dimensions

    full_ws = wb.create_sheet("full_audit")
    full_dicts = [asdict(row) for row in full_rows]
    full_columns = list(full_dicts[0]) if full_dicts else []
    write_sheet(full_ws, full_columns, full_dicts)
    full_ws.freeze_panes = "A2"
    full_ws.auto_filter.ref = full_ws.dimensions

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    wb.save(output_xlsx)


def write_sheet(ws, columns: Sequence[str], rows: Sequence[dict[str, object]]) -> None:
    """Write rows to an openpyxl worksheet with wrapped text."""
    from openpyxl.styles import Alignment

    ws.append(list(columns))
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_probe_csv(rows: Sequence[ProbeRow], output_csv: Path) -> None:
    """Write all probe columns to CSV."""
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(asdict(rows[0]).keys()) if rows else list(ProbeRow.__dataclass_fields__)
    with output_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def md_cell(value: object) -> str:
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", " ")


def write_probe_markdown(rows: Sequence[ProbeRow], output_md: Path, output_csv: Path) -> None:
    """Write a compact human-readable validation report."""
    output_md.parent.mkdir(parents=True, exist_ok=True)
    table_cols = [
        ("dataset", "dataset"),
        ("child_id", "child"),
        ("age_months", "age"),
        ("utterance_clean", "utterance"),
        ("recommended_word_count", "words"),
        ("recommended_morpheme_count", "morph_MOR_proxy"),
        ("morpheme_count_suffix_heuristic", "morph_suffix"),
        ("recommended_syllable_count", "syll_rec"),
        ("syllable_count_cmudict_or_syllables_pkg", "syll_CMU_pkg"),
        ("syllable_count_cmudict", "syll_CMU_only"),
        ("syllable_count_cmudict_g2p", "syll_CMU_G2P"),
        ("syllable_count_pyphen", "syll_pyphen"),
        ("syllable_count_syllables_pkg", "syll_pkg"),
        ("recommended_phoneme_count", "phon_rec"),
        ("phoneme_count_cmudict", "phon_CMU_only"),
        ("phoneme_count_cmudict_g2p", "phon_CMU_G2P"),
        ("g2p_fallback_word_count", "g2p_n"),
        ("raw_repetition_marker", "raw_rep"),
        ("mor_surface_mismatch", "mor_surface_mismatch"),
        ("quality_flags", "flags"),
    ]
    lines = [
        "# Utterance Measurement Validation Probe",
        "",
        f"CSV: `{output_csv}`",
        "",
        "Reference hierarchy:",
        "",
        "- Words: cleaned orthographic lexical tokens, punctuation excluded.",
        "- Morphemes: raw CHAT `%mor` tier when present; suffix heuristic shown as fallback/comparison.",
        "- Syllables: recommended counts use CMUdict for known words and the `syllables` package for OOV forms as written.",
        "- Phonemes: recommended counts use CMUdict ARPABET for known words and g2p-en for OOV forms as written.",
        "- CMU-only columns remain blank for OOV rows, but recommended syllable/phoneme columns are complete.",
        "- G2P-derived syllables are shown as a diagnostic only; G2P is retained mainly for phonemes.",
        "- `morph_MOR_proxy` is an MLU-style `%mor` proxy; `morph_suffix` is surface-string aligned and counts retraced repetitions.",
        "- `raw_rep` and `mor_surface_mismatch` flag cases where CHAT retracing/repetition markup makes MOR/MLU and surface-string counts diverge.",
        "",
        "Important context note: in this project, scoring context means the preceding caretaker utterance window "
        "(`context_k1`, `context_k2`, `context_k3`). Word-level/Levshina-style bits would distribute target "
        "utterance surprisal over the words of the current target utterance; the optional conditioning context is "
        "still preceding context, not future words.",
        "",
        "Method references:",
        "",
        "- TalkBank/CLAN: MLU is computed from `%mor` tiers by default; CLAN/MOR is the domain-native final reference.",
        "- CMU Pronouncing Dictionary: North American English ARPABET pronunciations; vowels carry stress digits.",
        "- The `syllables` package is used as the current OOV syllable fallback; g2p-en is used as the current OOV phoneme fallback.",
        "",
        "| " + " | ".join(label for _field, label in table_cols) + " |",
        "| " + " | ".join("---" for _field, _label in table_cols) + " |",
    ]
    for row in rows:
        row_dict = asdict(row)
        lines.append("| " + " | ".join(md_cell(row_dict[field]) for field, _label in table_cols) + " |")
    output_md.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_datasets(value: str) -> tuple[str, ...]:
    parts = tuple(part.strip() for part in value.split(",") if part.strip())
    if not parts:
        raise argparse.ArgumentTypeError("at least one dataset is required")
    return parts


def build_probe(
    *,
    scoring_root: Path,
    raw_root: Path,
    datasets: Sequence[str],
    sample_size: int,
    seed: int,
    min_words: int,
    output_csv: Path,
    output_md: Path,
    review_csv: Path,
    token_csv: Path,
    review_xlsx: Path,
) -> list[ProbeRow]:
    candidates = load_candidate_probe_rows(
        scoring_root=scoring_root,
        raw_root=raw_root,
        datasets=datasets,
        min_words=min_words,
    )
    selected = select_probe_rows(candidates, sample_size=sample_size, seed=seed)
    write_probe_csv(selected, output_csv)
    write_probe_markdown(selected, output_md, output_csv)
    write_review_outputs(
        selected,
        review_csv=review_csv,
        token_csv=token_csv,
        review_xlsx=review_xlsx,
    )
    return selected


def build_cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--scoring-root", type=Path, default=DEFAULT_SCORING_ROOT)
    parser.add_argument("--raw-root", type=Path, default=DEFAULT_RAW_ROOT)
    parser.add_argument("--datasets", type=parse_datasets, default=DEFAULT_DATASETS)
    parser.add_argument("--sample-size", type=int, default=50)
    parser.add_argument("--seed", type=int, default=20260602)
    parser.add_argument("--min-words", type=int, default=1)
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--output-md", type=Path, default=DEFAULT_OUTPUT_MD)
    parser.add_argument("--review-csv", type=Path, default=DEFAULT_REVIEW_CSV)
    parser.add_argument("--token-csv", type=Path, default=DEFAULT_TOKEN_CSV)
    parser.add_argument("--review-xlsx", type=Path, default=DEFAULT_REVIEW_XLSX)
    return parser


def main(argv: Sequence[str] | None = None) -> None:
    args = build_cli().parse_args(argv)
    rows = build_probe(
        scoring_root=args.scoring_root,
        raw_root=args.raw_root,
        datasets=args.datasets,
        sample_size=args.sample_size,
        seed=args.seed,
        min_words=args.min_words,
        output_csv=args.output_csv,
        output_md=args.output_md,
        review_csv=args.review_csv,
        token_csv=args.token_csv,
        review_xlsx=args.review_xlsx,
    )
    print(f"[OK] wrote {len(rows)} validation rows")
    print(f"[OK] CSV: {args.output_csv}")
    print(f"[OK] Markdown: {args.output_md}")
    print(f"[OK] Review CSV: {args.review_csv}")
    print(f"[OK] Token CSV: {args.token_csv}")
    print(f"[OK] Review XLSX: {args.review_xlsx}")


if __name__ == "__main__":
    main()
